#!/usr/bin/python3

import time
import datetime as dt
import pytz
from datetime import datetime

import sys
import logging
from logging import info
import glob
import os

import requests
import influxdb_client
from influxdb_client import InfluxDBClient
from influxdb_client import BucketsService, Bucket, PostBucketRequest
   
# openpyxl + csv modules
import openpyxl

# Création du logguer Chemin  A Adapter 
logging.basicConfig(filename='/home/airvb/grdf2influx/releve.log', level=logging.INFO, format='%(asctime)s %(message)s')

print("Début")
logging.info("Début...")



#______________________

# Variables à adapter #
#______________________

# Chemin du fichier excel d'extraction GRDF le plus récent 

inputExcelFile = max(glob.glob('/home/airvb/Téléchargements/Donnees_informatives_*.xlsx'))

# Url serveur influx
# url="http://localhost:8086" 
# Doivent déja être créés dans influxdb  
# orga = 
# bucket =    
# token = 


url="http://192.168.88.150:8086"
orga = "oNe"
bucket = "BBB-ESSAI"
token = "kEV_VEFOkmAEm2NIG2qQKmZc9Er-modcxG1mHMg5j5BoZa3ysrhKabjlJ6EVupZa60s1et-Ow7WawL2-98J9gQ=="

prix_fixe = 0.690 # prix par jour de l'abonnement
prix_kwh = 0.0833 # prix du kwh

#connexion vers influx
client = influxdb_client.InfluxDBClient(
  url=url,
  token=token,
  org=orga
)

print("Verification connection Influx...", end=" ")
try:
    client.query_api().query(f"from(bucket:\"{bucket}\") |> range(start: -1m) |> limit(n:1)", orga)
except Exception as e:
    # Erreur de connexion 
    if e.status == 401:
        logging.warning("WARNING Connexion influx non concluante. bucket: " + bucket + " Org:" + orga)
        print(f"LA connexion à INFLUX n'est pas concluante sur le '{bucket}' "
                        f"ou le token est obsolète.") 

        client.close()
        sys.exit()

print("Connexion Influx > ", end=" ")
logging.info("Connexion Influx OK")

influxdb_buckets_api = client.buckets_api()
influxdb_org_api = influxdb_client.OrganizationsApi(client)

# Verification org existe
orgs = influxdb_org_api.find_organizations()

for org in orgs:
    if org.name != orga:
        logging.warning("WARNING Interrogation- Organisation inconnue: --> ARRET " + orga)
        print("Interrogation- Organisation inconnue: " + orga)
        sys.exit()

#vérification bucket existe
bucket_objects = influxdb_buckets_api.find_buckets()

connected = False
for x in range ( len(bucket_objects._buckets)) :
    if bucket in bucket_objects._buckets[x].name:
        # logging.info('Interrogation- Bucket Ok :'+ bucket)
        connected = True
        break

if connected == False :
    logging.warning('WARNING ARRET Interrogation- Bucket inconnu :' +bucket)
    print("Interrogation- Bucket inconnu: " + bucket)
    sys.exit()

#WriteAPI 
from influxdb_client.client.write_api import SYNCHRONOUS
write_api = client.write_api(write_options=SYNCHRONOUS)

query_api = client.query_api()

print("Ok influx")
logging.info("Ok influx")

# chargement fichier excel 
newWorkbook = openpyxl.load_workbook(inputExcelFile)

# Sheet1 du fichier xcel 
firstWorksheet = newWorkbook.active

# efface col A
firstWorksheet.delete_cols(1) # colonne A en moins
# Le n° du pce
pce =firstWorksheet["D4"].internal_value
# son alias
pce_alias = firstWorksheet["D5"].internal_value

#  on enleve les lignes du début du fichier 
firstWorksheet.delete_rows(1,8)

# assigne le nom des nouvelles colonnes
firstWorksheet["B1"] = "start"
firstWorksheet["C1"] = "end"
firstWorksheet["I1"] = "prix_fixe" # 
firstWorksheet["J1"] = "prix_kwh" # 
firstWorksheet["K1"] = "cout" #
firstWorksheet["L1"] = "month" #
firstWorksheet["M1"] = "month_name" #
firstWorksheet["N1"] = "month_short" #
firstWorksheet["O1"] = "weekday_no"  #
firstWorksheet["P1"] = "weekday_name" 
firstWorksheet["Q1"] = "year" #
firstWorksheet["R1"] = "pce" #
firstWorksheet["S1"] = "pce_alias" #



# calcul du cout pour chaque jour 
print("Calcul coût")
logging.info("Calcul coût")

for row in firstWorksheet.iter_cols(min_row=2, min_col=5,max_row = firstWorksheet.max_row, max_col=5):
    for cell in row:
        #print(cell.value, end=" ") 
        result = round((cell.value * prix_kwh) + prix_fixe ,2)
      
        firstWorksheet.cell(cell.row, column=9).value = prix_fixe # prix fixe par jour
        firstWorksheet.cell(cell.row, column=10).value = prix_kwh # prix kwh
        firstWorksheet.cell(cell.row, column=11).value = result # cout par jour  


last_col_value = firstWorksheet.cell(firstWorksheet.max_row, column=1).value
print("Premiere date du fichier:" , str(firstWorksheet.cell(2, 1).value))

# on interroge le bucket pour retrouver la derniere date enregistrée

query= (f'from(bucket:\"{bucket}\") \
    |> range(start: -90d) \
    |> filter(fn: (r) => r["_measurement"] == "Mesuré") \
    |> filter(fn: (r) => r["_field"] == "start") \
    |> drop(columns:["year","weekday_name","weekday_no","pce","pce_alias","month","month_short","_measurement","type", "month_name"]) \
    |> max()')
result = query_api.query(org=orga, query=query)

if len(result) > 0: # pour eviter si new bucket
    lastdate = result[0].records[0].values['_time']
    print("Derniere date enregistrée ds influx:" ,lastdate.strftime("%d/%m/%Y"))
    print("Derniere date du fichier d'import : ", firstWorksheet.cell(firstWorksheet.max_row, column=1).value)
    logging.info("Derniere date enregistrée ds influx:" + lastdate.strftime('%d/%b/%Y'))
else:
    # une date par defaut si aucune data
    lastdate = datetime. strptime("01/01/2018", '%d/%m/%Y')
    lastdate = lastdate.replace(tzinfo=pytz.utc)

# envoi vers influx
logging.info("Debut vers influx")

i = 0 # flag nb de jours ajoutés 

for cell in firstWorksheet["A"]:
    if float(cell.row) > 2 : # on passe la ligne des titres + la premiere ligne ( T° systematiquement absente du fichier d'extraction )
        # on passe les lignes non mesuré
        if str(firstWorksheet.cell(cell.row, column=8).value) == "Mesuré" :
            # col A 
            # on modifie le format de la date pour etre accepté par influx
            # infludbd date format : 2022-12-04T11:22:17.000Z
                
            datestr =(cell.value)
            
            # dtfr date francaise 
            dtfr=datestr.split("/")
            # composante de la date 
            year = int(dtfr[2])
            month = int(dtfr [1])
            day = int(dtfr[0])
            heure = 1
            min = 0
            
            dtfr = dt.datetime(year, month, day, heure , min).strftime("%Y-%m-%dT%H:%M:%SZ")
            # pour pouvoir comparer
            date_time_obj = datetime. strptime(dtfr, '%Y-%m-%dT%H:%M:%SZ')
            date_time_obj = date_time_obj.replace(tzinfo=pytz.utc)
            
            # on n'ajoute que les nouvelles dates
            if date_time_obj > lastdate:
                i = i + 1 # compte le nombre de jour ajoutés
                if i == 1 :
                    logging.info("On ajoute à partir du:" +  dtfr)
                
                # modification de la date      
                cell.value = dtfr

                # pour recuperer les différentes composantes de la date
                compodate = datetime.strptime(dtfr,"%Y-%m-%dT%H:%M:%SZ" )    
                firstWorksheet.cell(cell.row, column=12).value = compodate.month # mois chiffre
                firstWorksheet.cell(cell.row, column=13).value =  compodate.strftime('%B') # mois JANVIER
                firstWorksheet.cell(cell.row, column=14).value =  compodate.strftime('%b') # mois JAN
                firstWorksheet.cell(cell.row, column=15).value =  compodate.strftime('%w') # weekday_no
                firstWorksheet.cell(cell.row, column=16).value =  compodate.strftime('%A') # weekday name
                firstWorksheet.cell(cell.row, column=17).value =  compodate.strftime('%Y') # YEAR
                
                # on ajoute pce et son alias
                firstWorksheet.cell(cell.row, column=18).value =  pce
                firstWorksheet.cell(cell.row, column=19).value =  pce_alias
                
                # Les différentes datas à envoyer ds influx
                payload = {
                    "measurement": str(firstWorksheet.cell(cell.row, column=8).value), #Qualification du relevé
                    "tags": {
                        "month": str(firstWorksheet.cell(cell.row, column=12).value),
                        "month_name": str(firstWorksheet.cell(cell.row, column=13).value),
                        "month_short": str(firstWorksheet.cell(cell.row, column=14).value),
                        "weekday_no": str(firstWorksheet.cell(cell.row, column=15).value),
                        "weekday_name": str(firstWorksheet.cell(cell.row, column=16).value),
                        "year": str(firstWorksheet.cell(cell.row, column=17).value),
                        "pce": str(firstWorksheet.cell(cell.row, column=18).value),
                        "pce_alias": str(firstWorksheet.cell(cell.row, column=19).value)
                    },
                    "fields": {
                        "start": firstWorksheet.cell(cell.row, column=2).value,
                        "end": firstWorksheet.cell(cell.row, column=3).value,
                        "Volume consommé (m3)": float(firstWorksheet.cell(cell.row, column=4).value),
                        "Energie consommée (kWh)": float(firstWorksheet.cell(cell.row, column=5).value),
                        "Coefficient de conversion": firstWorksheet.cell(cell.row, column=6).value,
                        "Température locale (°C)": float(firstWorksheet.cell(cell.row, column=7).value),                
                        "prix_fixe": firstWorksheet.cell(cell.row, column=9).value,
                        "prix_kwh": firstWorksheet.cell(cell.row, column=10).value,
                        "cout": firstWorksheet.cell(cell.row, column=11).value

                    },
                    "time": firstWorksheet.cell(cell.row, column=1).value #Date de relevé
                }

                # on ajoute ds la bd 
                payloads = []
                payloads.append(payload)
                write_api.write(bucket=bucket,org=orga,record=payloads) 
                # delai nécessaire pour influx
                time.sleep(0.3) 
                # print (payload)
# fin influx
write_api.close()

logging.info("Fin influx")

last_col_value = firstWorksheet.cell(firstWorksheet.max_row, column=1).value
logging.info("Derniere date retenue :" + str(last_col_value))
logging.info( str(i) + " jours ajoutés" )

print("Terminé ," , i , "jours ajoutés.")

